from django.http import HttpResponse
from django.utils import timezone
from django_filters.rest_framework import DjangoFilterBackend
from openpyxl import Workbook
from rest_framework import mixins, status, viewsets
from rest_framework.decorators import action
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from rest_framework.viewsets import GenericViewSet

from evis.attendees.api.serializers import AttendeesSerializer, CPDSerializer
from evis.attendees.models import CPD, Attendees
from evis.blog.models import Blog, News, Testimonials
from evis.conference.api.serializers import ContractFormAdminSerializer, ContractFormSerializer
from evis.conference.models import AboutConferenceImage, ContractFile, ContractForm
from evis.dashboard.models import Mail
from evis.dashboard.pagination import LargeResultsSetPagination, RegisterInterestPagination, SmallResultsSetPagination
from evis.dashboard.serializers.advisors import (
    AdvisorsCreateSerializer,
    AdvisorsDetailSerializer,
    AdvisorsListSerializer,
    AdvisorsUpdateSerializer,
)
from evis.dashboard.serializers.blogs import (
    BlogDetailSerializer,
    BlogListSerializer,
    BlogUpdateSerializer,
    NewsCreateSerializer,
    NewsDetailSerializer,
    NewsListSerializer,
    NewsUpdateSerializer,
    TestimonialsCreateSerializer,
    TestimonialsDetailSerializer,
    TestimonialsListSerializer,
    TestimonialsUpdateSerializer,
)
from evis.dashboard.serializers.conference import (
    ConferenceImageCreateSerializer,
    ConferenceImageSerializer,
    ContractFileListSerializer,
    ContractFileSerializer,
)
from evis.dashboard.serializers.exhibitors import ExhibitorVersionSerializer
from evis.dashboard.serializers.mail import MailSerializer
from evis.dashboard.serializers.speakers import (
    SpeakerCreateSerializer,
    SpeakerListSerializer,
    SpeakersVersionListSerializer,
    SpeakersVersionSerializer,
    SpeakerUpdateSerializer,
)
from evis.exhibitor.api.serializers import (
    ExhibitorCreateSerializer,
    ExhibitorDetailSerializer,
    ExhibitorListSerializer,
    ExhibitorUpdateSerializer,
    OrganizersSerializer,
    ParticipantSerializer,
    WhyExhibitSerializer,
    WhyToAttendEvisSerializer,
)
from evis.exhibitor.filters import ExhibitorFilter
from evis.exhibitor.models import Exhibitor, ExhibitorVersion, Organizers, Participant, WhyExhibit, WhyToAttendEvis
from evis.home.api.serializers import ContactSerializer, StatisticSerializer, TimerSerializer
from evis.home.models import Contact, Statistic, Timer
from evis.partners_sponsrs.api.serializers import PartnerAndSponserSerializer
from evis.partners_sponsrs.filters import PartnerSponserFilter
from evis.partners_sponsrs.models import PartnerAndSponser
from evis.speakers.api.serializers import SpeakerDetailSerializer
from evis.speakers.filters import SpeakersFilter
from evis.speakers.models import Advisors, Speakers, SpeakersVersion
from evis.users.api.serializers import RegisterLinkSerializer
from evis.users.models import RegisterLink
from evis.visit.api.serializers import (
    RegisterInterestListSerializer,
    RegisterInterestSerializer,
    SubscripeNewsSerializer,
)
from evis.visit.models import RegisterInterest, SubscripeNews

from .filters import ContractFormFilter
from .permissions import IsAdmin


# Speakers
class SpeakersVersionViewSet(viewsets.ModelViewSet):
    queryset = SpeakersVersion.objects.all()
    serializer_class = SpeakersVersionSerializer
    permission_classes = [IsAdmin]

    def get_serializer_class(self):
        if self.action == "list":
            return SpeakersVersionListSerializer
        return SpeakersVersionSerializer


class SpeakersViewSet(viewsets.ModelViewSet):
    queryset = Speakers.objects.prefetch_related("version").all()
    serializer_class = SpeakerCreateSerializer
    pagination_class = LargeResultsSetPagination
    filter_backends = [DjangoFilterBackend]
    filterset_class = SpeakersFilter
    permission_classes = [IsAdmin]

    def get_serializer_class(self):
        if self.action == "list":
            return SpeakerListSerializer
        elif self.action == "retrieve":
            return SpeakerDetailSerializer
        elif self.action == "update" or self.action == "partial_update":
            return SpeakerUpdateSerializer
        return SpeakerCreateSerializer


# Exhibitors


class ExhibitorVersionViewSet(viewsets.ModelViewSet):
    queryset = ExhibitorVersion.objects.all()
    serializer_class = ExhibitorVersionSerializer
    permission_classes = [IsAdmin]


class ExhibitorViewSet(viewsets.ModelViewSet):
    queryset = Exhibitor.objects.prefetch_related("version").all()
    serializer_class = ExhibitorDetailSerializer
    pagination_class = LargeResultsSetPagination
    filter_backends = [
        DjangoFilterBackend,
    ]
    filterset_class = ExhibitorFilter
    permission_classes = [IsAdmin]
    search_fields = ["name"]

    def get_serializer_class(self):
        if self.action == "list":
            return ExhibitorListSerializer
        elif self.action == "retrieve":
            return ExhibitorDetailSerializer
        elif self.action == "update" or self.action == "partial_update":
            return ExhibitorUpdateSerializer
        return ExhibitorCreateSerializer


class ParticipantViewSet(viewsets.ModelViewSet):
    queryset = Participant.objects.all()
    serializer_class = ParticipantSerializer
    pagination_class = SmallResultsSetPagination
    permission_classes = [IsAdmin]


class WhyExhibitViewSet(viewsets.ModelViewSet):
    queryset = WhyExhibit.objects.all()
    serializer_class = WhyExhibitSerializer
    permission_classes = [IsAdmin]


class WhyToAttendEvisViewSet(viewsets.ModelViewSet):
    queryset = WhyToAttendEvis.objects.all()
    serializer_class = WhyToAttendEvisSerializer
    permission_classes = [IsAdmin]


# Blogs


class BlogViewSet(viewsets.ModelViewSet):
    queryset = Blog.objects.all()
    serializer_class = BlogDetailSerializer
    pagination_class = LargeResultsSetPagination
    permission_classes = [IsAdmin]

    def get_serializer_class(self):
        if self.action == "list":
            return BlogListSerializer
        elif self.action == "update" or self.action == "partial_update":
            return BlogUpdateSerializer
        return BlogDetailSerializer


# News
class NewsViewSet(viewsets.ModelViewSet):
    queryset = News.objects.all()
    serializer_class = NewsDetailSerializer
    pagination_class = LargeResultsSetPagination
    permission_classes = [IsAdmin]

    def get_serializer_class(self):
        if self.action == "list":
            return NewsListSerializer
        elif self.action == "create":
            return NewsCreateSerializer
        elif self.action == "update" or self.action == "partial_update":
            return NewsUpdateSerializer
        return NewsDetailSerializer


# Testimonials


class TestimonialsViewSet(viewsets.ModelViewSet):
    queryset = Testimonials.objects.all()
    serializer_class = TestimonialsDetailSerializer
    pagination_class = LargeResultsSetPagination
    permission_classes = [IsAdmin]

    def get_serializer_class(self):
        if self.action == "list":
            return TestimonialsListSerializer
        elif self.action == "create":
            return TestimonialsCreateSerializer
        elif self.action == "update" or self.action == "partial_update":
            return TestimonialsUpdateSerializer
        return TestimonialsDetailSerializer


# Register Link


class RegisterLinkViewSet(viewsets.ModelViewSet):
    queryset = RegisterLink.objects.all()
    serializer_class = RegisterLinkSerializer
    permission_classes = [IsAdmin]


# Advisors Board


class AdvisorsViewSet(viewsets.ModelViewSet):
    queryset = Advisors.objects.all()
    serializer_class = AdvisorsDetailSerializer
    pagination_class = LargeResultsSetPagination
    permission_classes = [IsAdmin]

    def get_serializer_class(self):
        if self.action == "list":
            return AdvisorsListSerializer
        elif self.action == "create":
            return AdvisorsCreateSerializer
        elif self.action == "update" or self.action == "partial_update":
            return AdvisorsUpdateSerializer
        return AdvisorsDetailSerializer


# Events


class TimerViewSet(viewsets.ViewSet):
    permission_classes = [IsAdmin]
    serializer_class = TimerSerializer
    queryset = Timer.objects.all()

    def retrieve(self, request, pk=None):
        timer = Timer.get_solo()
        serializer = TimerSerializer(timer)
        return Response(serializer.data)

    def list(self, request):
        timer = Timer.get_solo()
        serializer = TimerSerializer(timer)
        return Response(serializer.data)

    def update(self, request, pk=None):
        timer = Timer.get_solo()
        serializer = TimerSerializer(timer, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def partial_update(self, request, pk=None):
        timer = Timer.get_solo()
        serializer = TimerSerializer(timer, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


# Statistics


class StatisticViewSet(viewsets.ViewSet):
    serializer_class = StatisticSerializer
    permission_classes = [IsAdmin]
    queryset = Statistic.objects.all()

    def retrieve(self, request, pk=None):
        statistic = Statistic.get_solo()
        serializer = StatisticSerializer(statistic)
        return Response(serializer.data)

    def list(self, request):
        statistic = Statistic.get_solo()
        serializer = StatisticSerializer(statistic)
        return Response(serializer.data)

    def update(self, request, pk=None):
        statistic = Statistic.get_solo()
        serializer = StatisticSerializer(statistic, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def partial_update(self, request, pk=None):
        statistic = Statistic.get_solo()
        serializer = StatisticSerializer(statistic, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


# Conferences


class AboutConferenceImageViewSet(viewsets.ModelViewSet):
    queryset = AboutConferenceImage.objects.all()
    serializer_class = ConferenceImageSerializer
    permission_classes = [IsAdmin]

    # delete all images
    @action(detail=False, methods=["delete"])
    def delete_all(self, request):
        AboutConferenceImage.objects.all().delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    def get_serializer_class(self):
        if (
            self.action == "list"
            or self.action == "update"
            or self.action == "partial_update"
            or self.action == "retrieve"
        ):
            return ConferenceImageSerializer
        return ConferenceImageCreateSerializer

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            self.perform_create(serializer)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=True)
        if serializer.is_valid():
            self.perform_update(serializer)
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class ContractFileViewSet(viewsets.ModelViewSet):
    serializer_class = ContractFileListSerializer
    queryset = ContractFile.objects.all()
    permission_classes = [IsAdmin]

    def get_serializer_class(self):
        if self.action == "list":
            return ContractFileListSerializer
        return ContractFileSerializer

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        self.perform_destroy(instance)
        return Response(status=status.HTTP_204_NO_CONTENT)


class ContractFileTypeViewSet(GenericViewSet):
    def list(self, request):
        contract_file_types = ContractFile.objects.values_list("type", flat=True).distinct()
        return Response(contract_file_types)


class RegisterInterestTitlesViewSet(GenericViewSet):
    def list(self, request):
        titles = RegisterInterest.objects.values_list("title", flat=True).distinct()
        return Response(titles)


# Contract Form


class ContractFormViewSet(viewsets.ModelViewSet):
    queryset = ContractForm.objects.all()
    serializer_class = ContractFormSerializer
    permission_classes = [IsAdmin]

    pagination_class = LargeResultsSetPagination
    filter_backends = [DjangoFilterBackend]
    filterset_class = ContractFormFilter

    allowed_methods = ["list", "retrieve"]

    def get_serializer_class(self):
        if self.action == "retrieve":
            return ContractFormAdminSerializer
        return ContractFormSerializer

    @action(detail=False, methods=["get"])
    def download_by_filter(self, request):
        """Download contract forms by filter

        Args:
            request (contract_file__type): (
                ("one", "One"),
                ("two", "Two"),
            )

        Returns:
            file: Download It as Excel File
        """
        contract_file_type = request.query_params.get("contract_file__type")

        if contract_file_type is None:
            contract_file_type = self.kwargs.get("contract_file__type")

        if contract_file_type is None or contract_file_type == "":
            return Response({"error": "add `contract_file__type` filter to URL"}, status=status.HTTP_400_BAD_REQUEST)

        contract_forms = self.queryset.filter(contract_file__type=contract_file_type)

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = 'attachment; filename="contract_forms.xlsx"'

        wb = Workbook()
        ws = wb.active

        headers = ["Name", "Company", "Email", "Country", "Phone", "Industry", "Interested In"]

        ws.append(headers)

        for contract_form in contract_forms:
            data_row = [
                contract_form.name,
                contract_form.company,
                contract_form.email,
                contract_form.country,
                contract_form.phone,
                contract_form.industry,
                contract_form.interested_in,
            ]
            ws.append(data_row)

        wb.save(response)
        return response

    # remove permissions from download_all_excel
    @action(detail=False, methods=["get"])
    def download_all_excel(self, request):
        contract_forms = self.queryset.all()
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = 'attachment; filename="contract_forms.xlsx"'

        wb = Workbook()
        ws = wb.active

        # Add headers
        headers = ["Name", "Company", "Email", "Country", "Phone", "Industry", "Interested In"]
        ws.append(headers)

        # Add data
        for contract_form in contract_forms:
            data_row = [
                contract_form.name,
                contract_form.company,
                contract_form.email,
                contract_form.country,
                contract_form.phone,
                contract_form.industry,
                contract_form.interested_in,
            ]
            ws.append(data_row)

        wb.save(response)
        return response

    @action(detail=True, methods=["get"])
    def download_excel(self, request, pk=None):
        contract_form = self.get_object()
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="{contract_form.name}.xlsx"'

        wb = Workbook()
        ws = wb.active

        # Add headers
        headers = ["Name", "Company", "Email", "Country", "Phone", "Industry", "Interested In"]
        ws.append(headers)

        # Add data
        data_row = [
            contract_form.name,
            contract_form.company,
            contract_form.email,
            contract_form.country,
            contract_form.phone,
            contract_form.industry,
            contract_form.interested_in,
        ]
        ws.append(data_row)

        wb.save(response)
        return response


# Contact


class ContactViewSet(viewsets.ModelViewSet):
    queryset = Contact.objects.all()
    serializer_class = ContactSerializer
    permission_classes = [IsAdmin]


# Organizers


class OrganizersViewSet(viewsets.ModelViewSet):
    queryset = Organizers.objects.all()
    serializer_class = OrganizersSerializer
    pagination_class = LargeResultsSetPagination
    permission_classes = [IsAdmin]


# SubscripeNews


class SubscripeNewsViewSet(mixins.ListModelMixin, viewsets.GenericViewSet):
    queryset = SubscripeNews.objects.all()
    serializer_class = SubscripeNewsSerializer
    pagination_class = LargeResultsSetPagination
    permission_classes = [IsAdmin]

    @action(detail=False, methods=["get"])
    def download_all_excel_for_current_month(self, request):
        subscripe_news = self.queryset.filter(created_at__month=timezone.now().month)
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = 'attachment; filename="subscripe_news.xlsx"'

        wb = Workbook()
        ws = wb.active

        # Add headers
        headers = ["First Name", "Last Name", "Email", "Created At"]
        ws.append(headers)

        # Add data
        for subscripe_new in subscripe_news:
            # Remove the timezone information from the created_at attribute
            subscripe_new.created_at = subscripe_new.created_at.replace(tzinfo=None)

            data_row = [
                subscripe_new.first_name,
                subscripe_new.last_name,
                subscripe_new.email,
                subscripe_new.created_at,
            ]
            ws.append(data_row)

        wb.save(response)
        return response


# RegisterInterest


class RegisterInterestViewSet(mixins.ListModelMixin, viewsets.GenericViewSet):
    queryset = RegisterInterest.objects.all()
    serializer_class = RegisterInterestSerializer
    pagination_class = RegisterInterestPagination
    permission_classes = [IsAdmin]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = [
        "job_title",
        "interested_in",
    ]

    def get_serializer_class(self):
        if self.action == "list":
            return RegisterInterestListSerializer
        return RegisterInterestSerializer

    @action(detail=False, methods=["get"])
    def download_all_excel_for_current_month(self, request):
        register_interests = self.queryset.filter(created_at__month=timezone.now().month)
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = 'attachment; filename="register_interests.xlsx"'

        wb = Workbook()
        ws = wb.active

        # Add headers
        headers = [
            "Email",
            "Name",
            "Interested In",
            "Job Title",
            "Company Name",
            "Address",
            "City",
            "Country",
            "Phone Number",
            "Business Nature",
            "Created At",
        ]
        ws.append(headers)

        # Add data
        for register_interest in register_interests:
            # Remove the timezone information from the created_at attribute
            register_interest.created_at = register_interest.created_at.replace(tzinfo=None)

            data_row = [
                register_interest.email,
                register_interest.name,
                register_interest.interested_in,
                register_interest.job_title,
                register_interest.company_name,
                register_interest.address,
                register_interest.city,
                register_interest.country,
                register_interest.phone_number,
                register_interest.business_nature,
                register_interest.created_at,
            ]
            ws.append(data_row)

        wb.save(response)
        return response

    @action(detail=False, methods=["get"])
    def download_all_excel_by_filter_interested_in(self, request):
        """Download register interests by filter interested_in,
            if not found interested_in return queryset for all register interests

        add `interested_in` filter to URL Like That: http://example.com/?interested_in=CEO
        """
        interested_in = request.query_params.get("interested_in")

        register_interests = self.queryset.filter(interested_in=interested_in)

        if interested_in is None or interested_in == "":
            register_interests = self.queryset.all()

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = 'attachment; filename="register_interests.xlsx"'

        wb = Workbook()
        ws = wb.active

        # Add headers
        headers = [
            "Email",
            "Name",
            "Interested In",
            "Job Title",
            "Company Name",
            "Address",
            "City",
            "Country",
            "Phone Number",
            "Business Nature",
            "Created At",
        ]
        ws.append(headers)

        # Add data
        for register_interest in register_interests:
            # Remove the timezone information from the created_at attribute
            register_interest.created_at = register_interest.created_at.replace(tzinfo=None)

            data_row = [
                register_interest.email,
                register_interest.name,
                register_interest.interested_in,
                register_interest.job_title,
                register_interest.company_name,
                register_interest.address,
                register_interest.city,
                register_interest.country,
                register_interest.phone_number,
                register_interest.business_nature,
                register_interest.created_at,
            ]
            ws.append(data_row)

        wb.save(response)
        return response


class MailViewSet(mixins.ListModelMixin, mixins.CreateModelMixin, viewsets.GenericViewSet):
    queryset = Mail.objects.all()
    serializer_class = MailSerializer
    pagination_class = LargeResultsSetPagination
    permission_classes = [IsAdmin]

    def get_permissions(self):
        if self.action == "create":
            permission_classes = [IsAdmin]
        else:
            permission_classes = [AllowAny]
        return [permission() for permission in permission_classes]

    @action(detail=False, methods=["get"])
    def download_all_excel_for_current_month(self, request):
        mails = self.queryset.filter(created_at__month=timezone.now().month)
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = 'attachment; filename="mails.xlsx"'

        wb = Workbook()
        ws = wb.active

        # Add headers
        headers = ["Email", "Created At"]
        ws.append(headers)

        # Add data
        for mail in mails:
            # Remove the timezone information from the created_at attribute
            mail.created_at = mail.created_at.replace(tzinfo=None)

            data_row = [
                mail.email,
                mail.created_at,
            ]
            ws.append(data_row)

        wb.save(response)
        return response


# PartnerAndSponser
# ----------------------------------------------------------------
class PartnerAndSponserViewSet(viewsets.ModelViewSet):
    queryset = PartnerAndSponser.objects.all()
    serializer_class = PartnerAndSponserSerializer
    filter_backends = (DjangoFilterBackend,)
    filterset_class = PartnerSponserFilter
    permission_classes = [IsAdmin]


# Attendees
# ----------------------------------------------------------------


class AttendeesViewSet(viewsets.ModelViewSet):
    queryset = Attendees.objects.all()
    serializer_class = AttendeesSerializer
    permission_classes = [IsAdmin]


class CPDViewSet(viewsets.ModelViewSet):
    queryset = CPD.objects.all()
    serializer_class = CPDSerializer
    permission_classes = [IsAdmin]
